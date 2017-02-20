from openpyxl import load_workbook

for i in range(31):
    print(i+1)

    wb = load_workbook(filename = str(i+1) + '.xlsx')
    ws = wb.active

    cols = "BFHJ"

    save = ''
    
    r = 7
    
    while True:
        if ws['B'+str(r)].value == None:
            break
        for c in cols:
            save += ''.join(str(ws[c+str(r)].value).lower().split()) + ' '
        save += '\n'
        r += 1

    savefile = open(str(i+1) + ".txt", 'w')
    savefile.write(save)
    savefile.close()