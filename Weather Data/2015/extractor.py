from openpyxl import load_workbook

wb = load_workbook(filename = '2015.xlsx')
ws = wb.active

cols = "HIJKL"

save = ''

for i in range(31):
    save += str(i+1) + ' '
    for c in cols:
        save += str(ws[c+str(102+i)].value) + ' '
    save += '\n'

savefile = open("weather-2015-Aug.txt", 'w')
savefile.write(save)
savefile.close()